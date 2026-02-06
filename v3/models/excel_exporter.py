"""
엑셀 및 PDF 출력 모듈
배합 기록을 엑셀 파일로 출력하고, 스캔 효과를 적용하여 PDF로 변환합니다.
"""
import os
import shutil
import warnings
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Border, Side, Alignment
from config.config_manager import config
from utils.logger import logger

# PDF 변환 및 이미지 처리를 위한 라이브러리
import win32com.client
import fitz  # PyMuPDF
from PIL import Image, ImageEnhance, ImageFilter
import numpy as np


class ExcelExporter:
    """엑셀 및 PDF 출력 클래스"""

    def __init__(self):
        """초기화"""
        base_path = config.get("paths.output", "실적서")
        self.excel_folder = os.path.join(base_path, "excel")
        self.pdf_folder = os.path.join(base_path, "pdf")

        for folder in [self.excel_folder, self.pdf_folder]:
            os.makedirs(folder, exist_ok=True)

        self.cell_mapping = config.get("excel.cell_mapping", {})
        self.template_file = os.path.join("resources", "template.xlsx")
        logger.debug(f"ExcelExporter 초기화: excel={self.excel_folder}, pdf={self.pdf_folder}")

    def export_to_excel(self, data, include_image=False, image_path=None, include_work_time=True):
        """배합 기록을 엑셀 파일로 출력"""
        try:
            if not os.path.exists(self.template_file):
                logger.warning(f"템플릿 파일이 없습니다: {self.template_file}")
                return None

            output_file = os.path.join(self.excel_folder, f"{data['product_lot']}.xlsx")
            shutil.copy2(self.template_file, output_file)

            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore",
                    message="wmf image format is not supported*",
                    category=UserWarning,
                    module="openpyxl",
                )
                wb = load_workbook(output_file)
            ws = wb.active

            # 데이터 입력
            self._fill_excel_data(ws, data, include_work_time)

            if include_image and image_path:
                if not self._add_image_to_worksheet(ws, image_path):
                    logger.warning("이미지 삽입에 실패했지만 엑셀 파일은 정상 생성됩니다.")

            data_end_row = self.cell_mapping.get('data_start_row', 7) + len(data.get('materials', [])) - 1
            self._format_worksheet(ws, data_end_row)

            wb.save(output_file)
            wb.close()
            logger.info(f"엑셀 파일 생성 완료: {output_file}")
            return output_file
        except Exception as e:
            logger.error(f"엑셀 출력 오류: {e}", exc_info=True)
            return None

    def export_to_pdf(self, excel_file, effects_params: dict):
        """
        엑셀 파일을 스캔 효과가 적용된 PDF로 변환합니다.
        (Excel -> Temp PDF -> Image -> Effects -> Final PDF)
        """
        temp_pdf_path = None
        effects_params = effects_params or {}
        try:
            if not excel_file or not os.path.exists(excel_file):
                logger.error(f"PDF로 변환할 엑셀 파일이 없습니다: {excel_file}")
                return None

            pdf_file = excel_file.replace('.xlsx', '.pdf')
            final_pdf_path = os.path.join(self.pdf_folder, os.path.basename(pdf_file))
            temp_pdf_path = os.path.join(self.excel_folder, f"temp_{os.path.basename(pdf_file)}")

            # 1. 엑셀 -> 임시 고품질 PDF
            self._excel_to_temp_pdf(excel_file, temp_pdf_path)

            # 2. 임시 PDF -> 이미지 목록
            page_images = self._pdf_to_images(temp_pdf_path, effects_params)
            if not page_images:
                raise ValueError("PDF를 이미지로 변환하는 데 실패했습니다.")

            # 3. 각 이미지에 스캔 효과 적용
            processed_images = [self._apply_scan_effects(img, effects_params) for img in page_images]

            # 4. 효과 적용된 이미지 목록 -> 최종 PDF
            self._images_to_final_pdf(processed_images, final_pdf_path)
            
            logger.info(f"스캔 효과 적용 PDF 생성 완료: {final_pdf_path}")
            return final_pdf_path

        except Exception as e:
            logger.error(f"PDF 변환 전체 프로세스 오류: {e}", exc_info=True)
            return None
        finally:
            # 5. 임시 파일 정리
            self._cleanup([temp_pdf_path])

    def _excel_to_temp_pdf(self, excel_path, pdf_path):
        """엑셀 파일을 PDF로 변환 (win32com 사용)"""
        excel = None
        workbook = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            abs_excel_path = os.path.abspath(excel_path)
            abs_pdf_path = os.path.abspath(pdf_path)
            
            workbook = excel.Workbooks.Open(abs_excel_path)
            workbook.Worksheets(1).ExportAsFixedFormat(0, abs_pdf_path)
            logger.debug(f"임시 PDF 생성 완료: {pdf_path}")
        finally:
            if workbook:
                workbook.Close(SaveChanges=False)
            if excel:
                excel.Quit()

    def _pdf_to_images(self, pdf_path, params: dict):
        """PDF를 Pillow 이미지 목록으로 변환 (PyMuPDF 사용)"""
        images = []
        doc = fitz.open(pdf_path)
        dpi = params.get("dpi", 250)
        for page in doc:
            pix = page.get_pixmap(dpi=dpi)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            images.append(img)
        doc.close()
        logger.debug(f"PDF를 이미지로 변환 완료: {len(images)} 페이지, DPI: {dpi}")
        return images

    def _apply_scan_effects(self, image, params):
        """이미지에 스캔 효과 적용"""
        blur = params.get("blur_radius", 0.3)
        noise = params.get("noise_range", 25)
        contrast = params.get("contrast_factor", 1.4)
        brightness = params.get("brightness_factor", 1.1)

        proc_img = image.filter(ImageFilter.GaussianBlur(radius=blur))
        
        img_array = np.array(proc_img)
        noise_array = np.random.randint(-noise, noise, img_array.shape, dtype='int16')
        img_array = np.clip(img_array + noise_array, 0, 255).astype('uint8')
        proc_img = Image.fromarray(img_array)

        enhancer = ImageEnhance.Contrast(proc_img)
        proc_img = enhancer.enhance(contrast)

        enhancer = ImageEnhance.Brightness(proc_img)
        proc_img = enhancer.enhance(brightness)
        
        return proc_img

    def _images_to_final_pdf(self, image_list, output_path):
        """이미지 목록을 최종 PDF로 저장"""
        if not image_list:
            raise ValueError("PDF로 저장할 이미지가 없습니다.")
        
        image_list[0].save(
            output_path, "PDF", resolution=100.0, save_all=True, append_images=image_list[1:]
        )

    def _cleanup(self, files):
        """임시 파일 삭제"""
        for f in files:
            if f and os.path.exists(f):
                try:
                    os.remove(f)
                    logger.debug(f"임시 파일 삭제: {f}")
                except Exception as e:
                    logger.warning(f"임시 파일 삭제 실패: {f}, 오류: {e}")

    def _fill_excel_data(self, ws, data, include_work_time=True):
        """엑셀 시트에 데이터 채우기"""
        if 'date' in self.cell_mapping: ws[self.cell_mapping['date']] = f"작업일: {data.get('work_date', '')}"
        if 'scale' in self.cell_mapping: ws[self.cell_mapping['scale']] = f"저울: {data.get('scale', '')}"
        if 'worker' in self.cell_mapping: ws[self.cell_mapping['worker']] = f"작업자 : {data.get('worker', '')}"
        
        if include_work_time:
            if 'work_time' in self.cell_mapping: ws[self.cell_mapping['work_time']] = f"작업시간 : {data.get('work_time', '')}"
        else:
            if 'work_time' in self.cell_mapping: ws[self.cell_mapping['work_time']] = ""
            
        if 'product_lot' in self.cell_mapping: ws[self.cell_mapping['product_lot']] = data.get('product_lot', '')
        if 'total_amount' in self.cell_mapping: ws[self.cell_mapping['total_amount']] = data.get('total_amount', 0) / 100

        start_row = self.cell_mapping.get('data_start_row', 7)
        for idx, material in enumerate(data.get('materials', [])):
            row = start_row + idx
            if 'material_name_col' in self.cell_mapping: ws[f"{self.cell_mapping['material_name_col']}{row}"] = material.get('material_name', '')
            if 'material_lot_col' in self.cell_mapping: ws[f"{self.cell_mapping['material_lot_col']}{row}"] = material.get('material_lot', '')
            if 'ratio_col' in self.cell_mapping: ws[f"{self.cell_mapping['ratio_col']}{row}"] = material.get('ratio', 0)
            if 'theory_amount_col' in self.cell_mapping: ws[f"{self.cell_mapping['theory_amount_col']}{row}"] = material.get('theory_amount', 0)
            if 'actual_amount_col' in self.cell_mapping: ws[f"{self.cell_mapping['actual_amount_col']}{row}"] = material.get('actual_amount', 0)

    def _format_worksheet(self, ws, data_end_row):
        """워크시트 서식 정리"""
        self._delete_empty_rows(ws, data_end_row)
        self._apply_cell_merges(ws, data_end_row)
        self._apply_borders(ws, data_end_row)

    def _add_image_to_worksheet(self, ws, image_path):
        """워크시트 G2 셀에 이미지 추가"""
        try:
            if not os.path.exists(image_path):
                logger.warning(f"이미지 파일이 존재하지 않습니다: {image_path}")
                return False
            img = OpenpyxlImage(image_path)
            img.width = 228
            img.height = 65
            img.anchor = "G2"
            ws.add_image(img)
            return True
        except Exception as e:
            logger.error(f"이미지 추가 중 오류: {e}", exc_info=True)
            return False

    def _delete_empty_rows(self, ws, data_end_row):
        """불필요한 빈 행 삭제"""
        try:
            for row_num in range(ws.max_row, data_end_row, -1):
                if all(ws.cell(row=row_num, column=col).value is None for col in range(1, 8)):
                    ws.delete_rows(row_num, 1)
        except Exception as e:
            logger.warning(f"행 삭제 중 오류: {e}")

    def _apply_cell_merges(self, ws, data_end_row):
        """셀 병합 적용"""
        try:
            ws.merge_cells(f'A6:A{data_end_row}')
            ws.merge_cells(f'B6:B{data_end_row}')
            ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B6'].alignment = Alignment(horizontal='center', vertical='center')
        except Exception as e:
            logger.warning(f"셀 병합 중 오류: {e}")

    def _apply_borders(self, ws, data_end_row):
        """테이블 경계선 적용"""
        try:
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in range(5, data_end_row + 1):
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        except Exception as e:
            logger.warning(f"경계선 적용 중 오류: {e}")
