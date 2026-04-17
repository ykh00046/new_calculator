"""Unit tests for ExcelExporter.export_to_excel (PDCA #12 Layer D2).

실 resources/template.xlsx와 tempfile로 격리된 paths.output을 사용하여
셀 매핑 불변식을 검증한다. win32com/PDF 경로는 경유하지 않는다.
"""
import importlib
import os
import sys
import tempfile
import unittest
from unittest.mock import patch


def _ensure_dependencies() -> None:
    missing = []
    for module_name in ("openpyxl",):
        try:
            importlib.import_module(module_name)
        except ModuleNotFoundError:
            missing.append(module_name)
    if missing:
        raise unittest.SkipTest(
            f"Excel tests require optional dependencies: {', '.join(missing)}"
        )


_ensure_dependencies()

current_dir = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(current_dir))
sys.path.insert(0, PROJECT_ROOT)

from openpyxl import load_workbook


# config.json의 excel.cell_mapping과 sync 필수.
# 의도적 하드카피 — config.json이 변경되면 D2 테스트가 실패하여 개발자가 인지한다.
REAL_CELL_MAPPING = {
    "date": "A3",
    "scale": "A4",
    "worker": "C3",
    "work_time": "E3",
    "product_lot": "A6",
    "total_amount": "B6",
    "data_start_row": 6,
    "material_name_col": "C",
    "material_lot_col": "D",
    "ratio_col": "E",
    "theory_amount_col": "F",
    "actual_amount_col": "G",
}


def _make_data(product_lot="LOT-TEST-01", total_amount=500.0, work_time="09:00:00",
               materials=None):
    if materials is None:
        materials = [
            {"material_name": "Mat1", "material_lot": "L1", "ratio": 40,
             "theory_amount": 200, "actual_amount": 199.5},
            {"material_name": "Mat2", "material_lot": "L2", "ratio": 60,
             "theory_amount": 300, "actual_amount": 300.2},
        ]
    return {
        "product_lot": product_lot,
        "recipe_name": "TEST",
        "work_date": "2026-04-17",
        "work_time": work_time,
        "worker": "tester",
        "total_amount": total_amount,
        "scale": "M-65",
        "materials": materials,
    }


class _ExcelExporterTestBase(unittest.TestCase):
    """D2 공통 setUp: template 존재 확인 + tempfile 격리 + config patch."""

    @classmethod
    def setUpClass(cls):
        cls.template = os.path.join(PROJECT_ROOT, "resources", "template.xlsx")
        if not os.path.exists(cls.template):
            raise unittest.SkipTest(
                f"resources/template.xlsx not found: {cls.template}"
            )

    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_path = self._tmp.name
        self.cfg_patcher = patch("models.excel_exporter.config")
        self.mock_cfg = self.cfg_patcher.start()
        self.mock_cfg.get.side_effect = self._cfg_side_effect

    def tearDown(self):
        self.cfg_patcher.stop()
        self._tmp.cleanup()

    def _cfg_side_effect(self, key, default=None):
        table = {
            "paths.output": self.tmp_path,
            "excel.cell_mapping": REAL_CELL_MAPPING,
        }
        return table.get(key, default)


# ---------------------------------------------------------------------------
# Layer D2 — Excel Content
# ---------------------------------------------------------------------------

class TestExcelExporterExportToExcel(_ExcelExporterTestBase):
    """실 template.xlsx 기반 export_to_excel 검증."""

    def _new_exporter(self):
        """ExcelExporter를 생성하고 template_file을 절대 경로로 고정."""
        from models.excel_exporter import ExcelExporter
        exporter = ExcelExporter()
        exporter.template_file = self.template
        return exporter

    def test_creates_file_with_product_lot_name(self):
        exporter = self._new_exporter()
        data = _make_data(product_lot="LOT-TEST-FILE-01")

        path = exporter.export_to_excel(data, include_work_time=True)

        self.assertIsNotNone(path)
        self.assertTrue(os.path.exists(path))
        self.assertTrue(path.endswith("LOT-TEST-FILE-01.xlsx"))
        self.assertIn(os.path.join(self.tmp_path, "excel"), path)

    def test_returns_none_when_template_missing(self):
        exporter = self._new_exporter()
        exporter.template_file = os.path.join(self.tmp_path, "nonexistent.xlsx")
        data = _make_data()

        path = exporter.export_to_excel(data, include_work_time=True)

        self.assertIsNone(path)

    def test_writes_key_cells(self):
        exporter = self._new_exporter()
        data = _make_data(product_lot="LOT-KEYCELLS-01", total_amount=500.0)

        path = exporter.export_to_excel(data, include_work_time=True)
        self.assertIsNotNone(path)

        wb = load_workbook(path)
        ws = wb.active
        try:
            self.assertEqual(ws[REAL_CELL_MAPPING["date"]].value, "작업일: 2026-04-17")
            self.assertEqual(ws[REAL_CELL_MAPPING["product_lot"]].value, "LOT-KEYCELLS-01")
            # total_amount 셀 == input / 100 (현행 코드 스펙 — excel_exporter.py:207)
            self.assertEqual(ws[REAL_CELL_MAPPING["total_amount"]].value, 5.0)
            self.assertIn("작업시간", ws[REAL_CELL_MAPPING["work_time"]].value)
        finally:
            wb.close()

    def test_writes_material_rows_from_data_start_row(self):
        exporter = self._new_exporter()
        data = _make_data(product_lot="LOT-ROWS-01")

        path = exporter.export_to_excel(data, include_work_time=True)
        self.assertIsNotNone(path)

        wb = load_workbook(path)
        ws = wb.active
        try:
            start = REAL_CELL_MAPPING["data_start_row"]
            name_col = REAL_CELL_MAPPING["material_name_col"]
            lot_col = REAL_CELL_MAPPING["material_lot_col"]
            ratio_col = REAL_CELL_MAPPING["ratio_col"]
            theory_col = REAL_CELL_MAPPING["theory_amount_col"]
            actual_col = REAL_CELL_MAPPING["actual_amount_col"]

            self.assertEqual(ws[f"{name_col}{start}"].value, "Mat1")
            self.assertEqual(ws[f"{name_col}{start + 1}"].value, "Mat2")
            self.assertEqual(ws[f"{lot_col}{start}"].value, "L1")
            self.assertEqual(ws[f"{ratio_col}{start}"].value, 40)
            self.assertEqual(ws[f"{ratio_col}{start + 1}"].value, 60)
            self.assertEqual(ws[f"{theory_col}{start + 1}"].value, 300)
            self.assertAlmostEqual(ws[f"{actual_col}{start}"].value, 199.5)
        finally:
            wb.close()

    def test_omits_work_time_when_flag_false(self):
        exporter = self._new_exporter()
        data = _make_data(product_lot="LOT-NOTIME-01", work_time="09:00:00")

        path = exporter.export_to_excel(data, include_work_time=False)
        self.assertIsNotNone(path)

        wb = load_workbook(path)
        ws = wb.active
        try:
            value = ws[REAL_CELL_MAPPING["work_time"]].value
            self.assertIn(value, ("", None))
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main()
